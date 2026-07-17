import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(name)s %(levelname)s %(message)s')
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pathlib import Path
from uuid import uuid4
from io import BytesIO
from openpyxl import Workbook, load_workbook
from .config import settings, get_api_key, get_base_url, get_model, set_runtime
from .models import *
from . import store
from .services import analyze, generate, add_version
app=FastAPI(title='Velocity Requirements API',version='2.0.0')
_cors_origins=[x.strip() for x in settings.cors_origins.split(',') if x.strip()]
# Always allow localhost variants so dev works regardless of .env
# for _o in ['http://localhost:5173','http://localhost:5174','http://127.0.0.1:5173','http://0.0.0.0:5173']:
#     if _o not in _cors_origins: _cors_origins.append(_o)
app.add_middleware(CORSMiddleware,allow_origins="*",allow_credentials=True, allow_methods=["*"],allow_headers=["*"])
@app.get('/api/config')
def get_config():
    key=get_api_key()
    return {
        'openai_api_key': ('*'*8+key[-4:]) if len(key)>8 else ('*'*len(key) if key else ''),
        'openai_base_url': get_base_url(),
        'openai_model': get_model(),
        'has_key': bool(key),
        'source': 'runtime' if key and key!=settings.openai_api_key else ('env' if settings.openai_api_key else 'none'),
    }
    

@app.put('/api/config')
async def put_config(req: LlmConfigRequest):
    # '__keep__' sentinel means: keep the existing key, only update url/model
    key=get_api_key() if req.openai_api_key.strip()=='__keep__' else req.openai_api_key
    set_runtime(key, req.openai_base_url, req.openai_model)
    # Quick connectivity probe
    import httpx as _httpx
    llm_mode='demo'; llm_error=None
    key=get_api_key()
    if key:
        try:
            async with _httpx.AsyncClient(timeout=8) as c:
                r=await c.post(
                    get_base_url().rstrip('/')+'/chat/completions',
                    headers={'Authorization':f'Bearer {key}','Content-Type':'application/json'},
                    json={'model':get_model(),'temperature':0,'max_tokens':1,'messages':[{'role':'user','content':'hi'}]}
                )
                llm_mode='configured' if r.status_code in (200,400) else 'error'
                if r.status_code not in (200,400): llm_error=f'HTTP {r.status_code}'
        except _httpx.ConnectError as e: llm_mode='unreachable'; llm_error=f'Cannot reach {get_base_url()}: {e}'
        except _httpx.TimeoutException: llm_mode='unreachable'; llm_error='Connection timed out'
        except Exception as e: llm_mode='error'; llm_error=str(e)
    return {'success':True,'llm_mode':llm_mode,'llm_error':llm_error}

@app.get('/api/health')
async def health():
    llm_mode='demo'
    llm_error=None
    if settings.openai_api_key:
        import httpx as _httpx
        try:
            async with _httpx.AsyncClient(timeout=8) as c:
                r=await c.post(
                    settings.openai_base_url.rstrip('/')+'/chat/completions',
                    headers={'Authorization':f'Bearer {settings.openai_api_key}','Content-Type':'application/json'},
                    json={'model':settings.openai_model,'temperature':0,'max_tokens':1,'messages':[{'role':'user','content':'hi'}]}
                )
                if r.status_code in (200,400): llm_mode='configured'
                else: llm_mode='error'; llm_error=f'HTTP {r.status_code}'
        except _httpx.ConnectError as e: llm_mode='unreachable'; llm_error=f'DNS/connect error: {e}'
        except _httpx.TimeoutException: llm_mode='unreachable'; llm_error='Connection timed out'
        except Exception as e: llm_mode='error'; llm_error=str(e)
    return {'status':'healthy','llm_mode':llm_mode,'llm_url':settings.openai_base_url,'llm_model':settings.openai_model,'llm_error':llm_error}
@app.get('/api/sessions')
def sessions(): return store.list_all()
@app.post('/api/sessions')
def create():
    s=Session(id=str(uuid4()),title='New requirements session'); store.save(s); return s
@app.get('/api/sessions/{sid}')
def get_session(sid:str):
    try:return store.load(sid)
    except FileNotFoundError:raise HTTPException(404,'Session not found')
@app.delete('/api/sessions/{sid}')
def delete_session(sid:str): store.delete(sid); return {'success':True}
@app.post('/api/sessions/{sid}/upload')
async def upload(sid:str,file:UploadFile=File(...)):
    s=get_session(sid); name=Path(file.filename or 'upload.bin').name; data=await file.read()
    if len(data)>settings.max_upload_mb*1024*1024: raise HTTPException(413,'File too large')
    out=settings.data_dir/'uploads'/f'{sid}_{uuid4().hex[:8]}_{name}'; out.write_bytes(data); s.files.append(str(out)); store.save(s)
    columns=[]
    if out.suffix.lower() in {'.xlsx','.xlsm'}:
        try:
            wb=load_workbook(out,read_only=True,data_only=True)
            ws=wb.active
            import re as _re
            # Strong requirement/engineering header keywords — high weight
            _strong_kw={
                'requirement','requirement statement','shall statement','description','req id','requirement id',
                'category','verification','verification method','acceptance criteria','acceptance','criteria',
                'source','rationale','origin','reference','owner','priority','status','comment','comments',
                'remarks','serial','s.no','id','method','test method','notes','result','pass fail',
                # DFMEA / FMEA specific
                'item','component','item component','function','function requirement',
                'potential failure mode','failure mode','failure effect','potential effects of failure',
                'severity','occurrence','detection','rpn','prevention','detection control',
                'recommended action','responsibility','target date','action result',
                'classification','special characteristic','cause','mechanism',
            }
            # Weak / metadata keywords that appear in non-header rows — penalise if these dominate
            _weak_kw={'internal','external','customer','confidential','draft','approved','open',
                      'closed','tbd','n/a','yes','no','high','medium','low','product','charger',
                      'scope','boundary','scope boundary','confidentiality'}
            best_cols=[]; best_score=-1
            for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row or 1,60),values_only=True):
                vals=[str(x).strip() for x in row if x not in (None,'')]
                if len(vals)<3: continue
                norm=[_re.sub(r'[^a-z0-9 /]+','',v.lower()).strip() for v in vals]
                strong_hits=sum(1 for n in norm if n in _strong_kw)
                weak_hits=sum(1 for n in norm if n in _weak_kw)
                # Score: strong hits weighted x5, column count bonus, penalise weak-only rows
                score=strong_hits*5 + len(vals) - weak_hits*3
                # Must have at least 1 strong hit to be considered a real header
                if strong_hits==0: continue
                if score>best_score:
                    best_score=score; best_cols=vals
            if len(best_cols)>=2: columns=best_cols
            wb.close()
        except Exception: pass
    return {'success':True,'file':name,'template_columns':columns}
@app.post('/api/sessions/{sid}/analyze')
async def do_analyze(sid:str,req:AnalyzeRequest):
    s=get_session(sid); s.stakeholder_needs=req.stakeholder_needs.strip(); s.title=(s.stakeholder_needs[:60] or 'Requirements session')
    context_parts=[p for p in [req.additional_context, req.clarification_answers] if p and p.strip()]
    text,questions=await analyze(s.stakeholder_needs, chr(10).join(context_parts)); s.analysis=text; s.clarification_questions=questions; s.messages += [{'role':'user','content':req.stakeholder_needs},{'role':'assistant','content':text}]
    data={'analysis':text,'clarification_questions':questions}
    if req.direct_generation:
        table=await generate(s.stakeholder_needs,s.analysis,req.template_columns); add_version(s,table,'generated'); data['table']=table.model_dump()
    store.save(s); return {'success':True,'stage':'clarification' if questions and not req.clarification_answers else 'analysis_complete','data':data}
@app.post('/api/sessions/{sid}/generate')
async def do_generate(sid:str,req:GenerateRequest):
    s=get_session(sid)
    if not s.analysis: raise HTTPException(400,'Run analysis first')
    table=await generate(s.stakeholder_needs,s.analysis,req.template_columns); v=add_version(s,table,'generated'); store.save(s); return {'success':True,'stage':'complete','data':{'table':table,'version':v.version}}
@app.put('/api/sessions/{sid}/table')
def save_table(sid:str,req:TableRequest):
    s=get_session(sid); table=RequirementTable(columns=req.columns,rows=req.rows); v=add_version(s,table,req.source); store.save(s); return {'success':True,'data':{'version':v.version}}
@app.post('/api/sessions/{sid}/actions/{action}')
async def action(sid:str,action:str,req:ActionRequest):
    s=get_session(sid)
    if action not in {'query','review','update','revise'}: raise HTTPException(404,'Unknown action')
    table=s.versions[s.active_version].table if s.active_version>=0 else None
    if action=='revise' and table:
        rows=[list(r) for r in table.rows]; ci=next((i for i,c in enumerate(table.columns) if 'comment' in c.name.lower()),None)
        if ci is not None:
            for r in rows:
                existing=str(r[ci]).strip(' |') if r[ci] else ''
                r[ci]=(existing+' | Revision: '+req.text).strip(' |') if existing else 'Revision: '+req.text
        nv=RequirementTable(title=table.title,columns=table.columns,rows=rows); v=add_version(s,nv,'revised'); store.save(s); return {'success':True,'data':{'response':'Revision captured in a new version.','table':nv,'version':v.version}}
    if action=='revise' and not table: raise HTTPException(400,'Generate requirements first')
    row_count=len(table.rows) if table else 0
    response=f"{action.title()} assessment: {req.text}. " + (f'The current table contains {row_count} requirements. ' if table else '') + 'Review traceability, measurability, feasibility, and verification impact before approval.'
    s.messages.append({'role':'assistant','content':response}); store.save(s); return {'success':True,'data':{'response':response}}
@app.get('/api/sessions/{sid}/diagram')
def diagram(sid:str):
    s=get_session(sid)
    if s.active_version<0: raise HTTPException(400,'Generate requirements first')
    t=s.versions[s.active_version].table; lines=['flowchart LR','A[Stakeholder Needs] --> B[Analysis]','B --> C[Requirements]']
    for i,row in enumerate(t.rows[:10]): lines.append(f'C --> R{i}[{str(row[0] if row else i+1).replace("[","(").replace("]",")")}]')
    return {'success':True,'data':{'mermaid':chr(10).join(lines)}}
@app.get('/api/sessions/{sid}/export.xlsx')
def export_xlsx(sid:str):
    s=get_session(sid)
    if s.active_version<0: raise HTTPException(400,'Generate requirements first')
    t=s.versions[s.active_version].table; wb=Workbook(); ws=wb.active; ws.title='Requirements'; ws.append([c.name for c in t.columns])
    for r in t.rows: ws.append(r)
    for cell in ws[1]: cell.font=cell.font.copy(bold=True)
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=requirements.xlsx'})

import logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(name)s %(levelname)s %(message)s",
)

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pathlib import Path
from uuid import uuid4
from io import BytesIO

from openpyxl import Workbook, load_workbook

from .config import settings, get_api_key, get_base_url, get_model, set_runtime
from .models import (
    Session, AnalyzeRequest, GenerateRequest,
    TableRequest, ActionRequest, LlmConfigRequest,
)
from . import store
from .services import analyze, generate, generate_diagram_data, run_action, add_version

app = FastAPI(title="Velocity Risk Assessor API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins="*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Config & health
# ---------------------------------------------------------------------------

@app.get("/api/config")
def get_config():
    key = get_api_key()
    return {
        "openai_api_key": ("*" * 8 + key[-4:]) if len(key) > 8 else ("*" * len(key) if key else ""),
        "openai_base_url": get_base_url(),
        "openai_model": get_model(),
        "has_key": bool(key),
        "source": "runtime" if key and key != settings.openai_api_key else ("env" if settings.openai_api_key else "none"),
    }


@app.put("/api/config")
async def put_config(req: LlmConfigRequest):
    key = get_api_key() if req.openai_api_key.strip() == "__keep__" else req.openai_api_key
    set_runtime(key, req.openai_base_url, req.openai_model)
    import httpx as _httpx
    llm_mode = "demo"
    llm_error = None
    key = get_api_key()
    if key:
        try:
            async with _httpx.AsyncClient(timeout=8) as c:
                r = await c.post(
                    get_base_url().rstrip("/") + "/chat/completions",
                    headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                    json={"model": get_model(), "temperature": 0, "max_tokens": 1,
                          "messages": [{"role": "user", "content": "hi"}]},
                )
                llm_mode = "configured" if r.status_code in (200, 400) else "error"
                if r.status_code not in (200, 400):
                    llm_error = f"HTTP {r.status_code}"
        except _httpx.ConnectError as e:
            llm_mode = "unreachable"
            llm_error = f"Cannot reach {get_base_url()}: {e}"
        except _httpx.TimeoutException:
            llm_mode = "unreachable"
            llm_error = "Connection timed out"
        except Exception as e:
            llm_mode = "error"
            llm_error = str(e)
    return {"success": True, "llm_mode": llm_mode, "llm_error": llm_error}


@app.get("/api/health")
async def health():
    llm_mode = "demo"
    llm_error = None
    if settings.openai_api_key:
        import httpx as _httpx
        try:
            async with _httpx.AsyncClient(timeout=8) as c:
                r = await c.post(
                    settings.openai_base_url.rstrip("/") + "/chat/completions",
                    headers={"Authorization": f"Bearer {settings.openai_api_key}",
                             "Content-Type": "application/json"},
                    json={"model": settings.openai_model, "temperature": 0, "max_tokens": 1,
                          "messages": [{"role": "user", "content": "hi"}]},
                )
                if r.status_code in (200, 400):
                    llm_mode = "configured"
                else:
                    llm_mode = "error"
                    llm_error = f"HTTP {r.status_code}"
        except _httpx.ConnectError as e:
            llm_mode = "unreachable"
            llm_error = f"DNS/connect error: {e}"
        except _httpx.TimeoutException:
            llm_mode = "unreachable"
            llm_error = "Connection timed out"
        except Exception as e:
            llm_mode = "error"
            llm_error = str(e)
    return {
        "status": "healthy",
        "llm_mode": llm_mode,
        "llm_url": settings.openai_base_url,
        "llm_model": settings.openai_model,
        "llm_error": llm_error,
    }


# ---------------------------------------------------------------------------
# Session CRUD
# ---------------------------------------------------------------------------

@app.get("/api/sessions")
def sessions():
    return store.list_all()


@app.post("/api/sessions")
def create():
    s = Session(id=str(uuid4()), title="New risk assessment session")
    store.save(s)
    return s


@app.get("/api/sessions/{sid}")
def get_session(sid: str):
    try:
        return store.load(sid)
    except FileNotFoundError:
        raise HTTPException(404, "Session not found")


@app.delete("/api/sessions/{sid}")
def delete_session(sid: str):
    store.delete(sid)
    return {"success": True}


# ---------------------------------------------------------------------------
# File upload — detects Excel template columns
# ---------------------------------------------------------------------------

@app.post("/api/sessions/{sid}/upload")
async def upload(sid: str, file: UploadFile = File(...)):
    s = get_session(sid)
    name = Path(file.filename or "upload.bin").name
    data = await file.read()
    if len(data) > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(413, "File too large")

    out = settings.data_dir / "uploads" / f"{sid}_{uuid4().hex[:8]}_{name}"
    out.write_bytes(data)
    s.files.append(str(out))
    store.save(s)

    columns: list[str] = []
    if out.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import re as _re
            wb = load_workbook(out, read_only=True, data_only=True)
            ws = wb.active

            _strong_kw = {
                # Risk register keywords
                "risk id", "risk statement", "risk description", "description",
                "category", "cause", "event", "impact", "likelihood", "probability",
                "severity", "occurrence", "detection", "rpn",
                "overall rating", "risk rating", "proposed mitigation", "mitigation",
                "treatment", "owner", "risk owner", "due date", "target date",
                "residual risk", "status", "existing controls", "controls",
                "control effectiveness", "contingency", "evidence", "source",
                "reference", "confidence", "notes", "remarks", "comments",
                "affected asset", "asset", "process",
                # DFMEA keywords
                "item", "component", "item component", "function",
                "potential failure mode", "failure mode", "failure effect",
                "potential effects of failure", "prevention controls",
                "detection controls", "recommended action", "responsibility",
                "action result", "classification", "special characteristic",
                "potential causes", "mechanism",
            }
            _weak_kw = {
                "internal", "external", "customer", "confidential", "draft",
                "approved", "open", "closed", "tbd", "n/a", "yes", "no",
                "high", "medium", "low", "product", "scope", "boundary",
            }

            best_cols: list[str] = []
            best_score = -1
            for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row or 1, 60), values_only=True
            ):
                vals = [str(x).strip() for x in row if x not in (None, "")]
                if len(vals) < 2:
                    continue
                norm = [_re.sub(r"[^a-z0-9 /]+", "", v.lower()).strip() for v in vals]
                strong_hits = sum(1 for n in norm if n in _strong_kw)
                weak_hits = sum(1 for n in norm if n in _weak_kw)
                score = strong_hits * 5 + len(vals) - weak_hits * 3
                if strong_hits == 0:
                    continue
                if score > best_score:
                    best_score = score
                    best_cols = vals
            if len(best_cols) >= 2:
                columns = best_cols
            wb.close()
        except Exception:
            pass

    return {"success": True, "file": name, "template_columns": columns}


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

@app.post("/api/sessions/{sid}/analyze")
async def do_analyze(sid: str, req: AnalyzeRequest):
    s = get_session(sid)
    s.stakeholder_needs = req.stakeholder_needs.strip()
    s.title = s.stakeholder_needs[:60] or "Risk assessment session"

    context_parts = [
        p for p in [req.additional_context, req.clarification_answers]
        if p and p.strip()
    ]
    text, questions = await analyze(
        s.stakeholder_needs, "\n".join(context_parts)
    )
    s.analysis = text
    s.clarification_questions = questions
    s.messages += [
        {"role": "user", "content": req.stakeholder_needs},
        {"role": "assistant", "content": text},
    ]

    data: dict = {"analysis": text, "clarification_questions": questions}

    if req.direct_generation:
        table = await generate(s.stakeholder_needs, s.analysis, req.template_columns)
        add_version(s, table, "generated")
        data["table"] = table.model_dump()

    store.save(s)
    stage = "clarification" if questions and not req.clarification_answers else "analysis_complete"
    return {"success": True, "stage": stage, "data": data}


# ---------------------------------------------------------------------------
# Generate risk register
# ---------------------------------------------------------------------------

@app.post("/api/sessions/{sid}/generate")
async def do_generate(sid: str, req: GenerateRequest):
    s = get_session(sid)
    if not s.analysis:
        raise HTTPException(400, "Run analysis first")
    table = await generate(s.stakeholder_needs, s.analysis, req.template_columns)
    v = add_version(s, table, "generated")
    store.save(s)
    return {"success": True, "stage": "complete", "data": {"table": table, "version": v.version}}


# ---------------------------------------------------------------------------
# Save / update table
# ---------------------------------------------------------------------------

@app.put("/api/sessions/{sid}/table")
def save_table(sid: str, req: TableRequest):
    from .models import RiskTable
    s = get_session(sid)
    table = RiskTable(columns=req.columns, rows=req.rows)
    v = add_version(s, table, req.source)
    store.save(s)
    return {"success": True, "data": {"version": v.version}}


# ---------------------------------------------------------------------------
# Actions: query / review / update / revise
# ---------------------------------------------------------------------------

@app.post("/api/sessions/{sid}/actions/{action}")
async def action(sid: str, action: str, req: ActionRequest):
    if action not in {"query", "review", "update", "revise"}:
        raise HTTPException(404, "Unknown action")
    s = get_session(sid)
    table = s.versions[s.active_version].table if s.active_version >= 0 else None

    if action == "revise" and table:
        from .models import RiskTable
        rows = [list(r) for r in table.rows]
        ci = next(
            (i for i, c in enumerate(table.columns) if "comment" in c.name.lower() or "notes" in c.name.lower()),
            None,
        )
        if ci is not None:
            for r in rows:
                existing = str(r[ci]).strip(" |") if r[ci] else ""
                r[ci] = (
                    (existing + " | Revision: " + req.text).strip(" |")
                    if existing
                    else "Revision: " + req.text
                )
        nv = RiskTable(title=table.title, columns=table.columns, rows=rows)
        v = add_version(s, nv, "revised")
        store.save(s)
        return {
            "success": True,
            "data": {
                "response": "Revision captured in a new version.",
                "table": nv,
                "version": v.version,
            },
        }

    if action == "revise" and not table:
        raise HTTPException(400, "Generate a risk register first")

    response = await run_action(action, req.text, table, s.analysis)
    s.messages.append({"role": "assistant", "content": response})
    store.save(s)
    return {"success": True, "data": {"response": response}}


# ---------------------------------------------------------------------------
# Diagram / heatmap
# ---------------------------------------------------------------------------

@app.get("/api/sessions/{sid}/diagram")
async def diagram(sid: str):
    s = get_session(sid)
    if s.active_version < 0:
        raise HTTPException(400, "Generate a risk register first")
    t = s.versions[s.active_version].table
    data = await generate_diagram_data(t)
    return {"success": True, "data": data}


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.get("/api/sessions/{sid}/export.xlsx")
def export_xlsx(sid: str):
    s = get_session(sid)
    if s.active_version < 0:
        raise HTTPException(400, "Generate a risk register first")
    t = s.versions[s.active_version].table
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"
    ws.append([c.name for c in t.columns])
    for r in t.rows:
        ws.append(r)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=risk_register.xlsx"},
    )

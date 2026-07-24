"""
common/backend/artifact_builder.py
────────────────────────────────────
Shared artifact-building utilities for all Velocity AI applications.

Provides:
  - build_xlsx_export()        : Write a table (columns + rows) to an .xlsx
                                  file in-memory and return a FastAPI
                                  StreamingResponse ready for download.
  - build_html_table_preview() : Render a list-of-dicts or columns+rows as
                                  a simple styled HTML table string.
  - build_findings_html()      : Render a list of finding dicts as a styled
                                  HTML findings/checklist table.
  - export_xlsx_response()     : One-liner helper used in /export.xlsx routes.

Usage example in an app's main.py:

    from common.backend.artifact_builder import export_xlsx_response

    @app.get("/api/sessions/{sid}/export.xlsx")
    def export_xlsx(sid: str):
        s = get_session(sid)
        if s.active_version < 0:
            raise HTTPException(400, "Generate data first")
        t = s.versions[s.active_version].table
        return export_xlsx_response(
            columns=[c.name for c in t.columns],
            rows=t.rows,
            sheet_title="Requirements",
            filename="requirements.xlsx",
        )
"""
from __future__ import annotations

import html as _html
from io import BytesIO
from typing import Any

from fastapi.responses import StreamingResponse


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def build_xlsx_bytes(
    columns: list[str],
    rows: list[list[Any]],
    sheet_title: str = "Sheet1",
    bold_header: bool = True,
    header_fill_color: str | None = "0D3B55",
    header_font_color: str | None = "FFFFFF",
) -> bytes:
    """
    Build an in-memory .xlsx file and return its raw bytes.

    Parameters
    ----------
    columns           : List of column header strings.
    rows              : List of row value lists.
    sheet_title       : Name of the worksheet.
    bold_header       : Make the header row bold.
    header_fill_color : Hex fill colour for the header row (None = no fill).
    header_font_color : Hex font colour for the header row (None = default).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise ImportError("openpyxl is required for XLSX export.") from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(columns)

    # Style header row
    if bold_header or header_fill_color or header_font_color:
        fill = (
            PatternFill("solid", fgColor=header_fill_color)
            if header_fill_color
            else None
        )
        font_kwargs: dict = {"bold": bold_header}
        if header_font_color:
            font_kwargs["color"] = header_font_color
        font = Font(**font_kwargs)
        for cell in ws[1]:
            if bold_header or header_font_color:
                cell.font = font
            if fill:
                cell.fill = fill

    for row in rows:
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def export_xlsx_response(
    columns: list[str],
    rows: list[list[Any]],
    sheet_title: str = "Sheet1",
    filename: str = "export.xlsx",
    bold_header: bool = True,
    header_fill_color: str | None = "0D3B55",
    header_font_color: str | None = "FFFFFF",
) -> StreamingResponse:
    """
    Build an .xlsx file and return a FastAPI StreamingResponse for download.

    This is the one-liner used in all /export.xlsx route handlers.
    """
    data = build_xlsx_bytes(
        columns=columns,
        rows=rows,
        sheet_title=sheet_title,
        bold_header=bold_header,
        header_fill_color=header_fill_color,
        header_font_color=header_font_color,
    )
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# HTML table preview
# ---------------------------------------------------------------------------

def build_html_table_preview(
    columns: list[str],
    rows: list[list[Any]],
    title: str = "",
    max_rows: int = 200,
    header_bg: str = "#0d3b55",
    header_color: str = "white",
    stripe_bg: str = "#f8fafc",
) -> str:
    """
    Render a columns + rows table as a styled HTML string.

    Suitable for embedding in a report preview or returning as a response body.
    """
    esc = _html.escape
    header_cells = "".join(f"<th>{esc(str(c))}</th>" for c in columns)
    body_rows = ""
    for i, row in enumerate(rows[:max_rows]):
        bg = stripe_bg if i % 2 == 1 else "white"
        cells = "".join(f"<td>{esc(str(v) if v is not None else '')}</td>" for v in row)
        body_rows += f"<tr style='background:{bg}'>{cells}</tr>\n"

    title_html = f"<h2 style='font-family:Inter,Arial;color:#0d3b55'>{esc(title)}</h2>" if title else ""

    return (
        f"{title_html}"
        f"<div style='overflow:auto'>"
        f"<table style='border-collapse:collapse;width:100%;font-family:Inter,Arial;font-size:13px;'>"
        f"<thead><tr style='background:{header_bg};color:{header_color};'>{header_cells}</tr></thead>"
        f"<tbody>{body_rows}</tbody>"
        f"</table></div>"
    )


# ---------------------------------------------------------------------------
# Findings / checklist HTML
# ---------------------------------------------------------------------------

# Severity → (bg colour, text colour)
_SEVERITY_COLORS: dict[str, tuple[str, str]] = {
    "critical": ("#fde8e8", "#b00020"),
    "error":    ("#fce4d6", "#d84315"),
    "warning":  ("#fff8e1", "#b26a00"),
    "info":     ("#e3f2fd", "#1565c0"),
}


def _severity_badge(sev: str) -> str:
    """Return an inline-styled HTML badge span for a severity level."""
    sev_lower = sev.lower()
    bg, color = _SEVERITY_COLORS.get(sev_lower, ("#eef3f7", "#18212b"))
    return (
        f"<span style=\""
        f"display:inline-block;padding:3px 10px;border-radius:20px;"
        f"font-size:11px;font-weight:600;"
        f"background:{bg};color:{color};"
        f"\">{_html.escape(sev)}</span>"
    )


def build_findings_html(
    findings: list[dict[str, Any]],
    title: str = "Findings",
    columns: list[str] | None = None,
) -> str:
    """
    Render a list of finding dicts as a styled HTML table.

    Each finding dict may contain: id, severity, category, text, recommendation,
    checked, description, and any other keys.

    Parameters
    ----------
    findings : List of finding dicts from check_results or report_checklist.
    title    : Section heading above the table.
    columns  : Column names to render. Defaults to
               ['ID', 'Severity', 'Category', 'Finding', 'Recommendation'].
    """
    esc = _html.escape
    cols = columns or ["ID", "Severity", "Category", "Finding", "Recommendation"]
    header_cells = "".join(
        f"<th style='background:#0d3b55;color:white;padding:10px;text-align:left;'>"
        f"{esc(c)}</th>"
        for c in cols
    )

    body_rows = ""
    for i, f in enumerate(findings):
        if not isinstance(f, dict):
            continue
        bg = "#f8fafc" if i % 2 == 1 else "white"
        fid       = esc(str(f.get("id") or f"F{i + 1}"))
        sev       = str(f.get("severity") or "info")
        category  = esc(str(f.get("category") or "—"))
        text      = esc(str(f.get("text") or f.get("description") or "—"))
        rec       = esc(str(f.get("recommendation") or "—"))

        row_data = {
            "ID":             fid,
            "Severity":       _severity_badge(sev),
            "Category":       category,
            "Finding":        text,
            "Recommendation": rec,
        }

        cells = ""
        for col in cols:
            val = row_data.get(col, "")
            cells += (
                f"<td style='border-bottom:1px solid #e2e8f0;"
                f"padding:8px 10px;vertical-align:top;background:{bg};'>"
                f"{val}</td>"
            )
        body_rows += f"<tr>{cells}</tr>\n"

    title_html = (
        f"<h2 style='font-family:Inter,Arial;color:#0d3b55;'>{esc(title)}</h2>"
        if title
        else ""
    )

    return (
        f"<div style='font-family:Inter,Arial;font-size:13px;'>"
        f"{title_html}"
        f"<div style='overflow:auto'>"
        f"<table style='border-collapse:collapse;width:100%;'>"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{body_rows}</tbody>"
        f"</table></div></div>"
    )

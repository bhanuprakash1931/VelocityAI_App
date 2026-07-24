"""
common/backend/template_handler.py
───────────────────────────────────
Shared Excel template column-detection and upload handler used by all
Velocity AI applications.

Provides:
  - detect_template_columns()  : Scan an uploaded Excel file for the best
                                  header row and return the column names.
  - save_upload()              : Save a raw uploaded file to the data/uploads
                                  directory with a session-scoped filename.
  - build_strong_kw_set()      : Returns the combined keyword set used for
                                  scoring header rows (Risk + Requirements +
                                  DFMEA keywords all in one place).

Usage in an app's main.py upload endpoint:

    from common.backend.template_handler import save_upload, detect_template_columns
    from .config import settings

    @app.post("/api/sessions/{sid}/upload")
    async def upload(sid: str, file: UploadFile = File(...)):
        out, name = await save_upload(
            file=file,
            sid=sid,
            uploads_dir=settings.data_dir / "uploads",
            max_upload_mb=settings.max_upload_mb,
        )
        columns = detect_template_columns(out) if out.suffix.lower() in {".xlsx", ".xlsm"} else []
        return {"success": True, "file": name, "template_columns": columns}
"""
from __future__ import annotations

import re
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException, UploadFile


# ---------------------------------------------------------------------------
# Keyword sets used for header-row scoring
# ---------------------------------------------------------------------------

# High-signal keywords — present in real header rows
_STRONG_KW: frozenset[str] = frozenset({
    # Requirements / general engineering
    "requirement", "requirement statement", "shall statement", "description",
    "req id", "requirement id", "category", "verification", "verification method",
    "acceptance criteria", "acceptance", "criteria", "source", "rationale",
    "origin", "reference", "owner", "priority", "status", "comment", "comments",
    "remarks", "serial", "s.no", "id", "method", "test method", "notes",
    "result", "pass fail",
    # Risk register
    "risk id", "risk statement", "risk description", "cause", "event",
    "impact", "likelihood", "probability", "severity", "occurrence", "detection",
    "rpn", "overall rating", "risk rating", "proposed mitigation", "mitigation",
    "treatment", "risk owner", "due date", "target date", "residual risk",
    "existing controls", "controls", "control effectiveness", "contingency",
    "evidence", "confidence", "affected asset", "asset", "process",
    # DFMEA / FMEA
    "item", "component", "item component", "function", "function requirement",
    "potential failure mode", "failure mode", "failure effect",
    "potential effects of failure", "prevention controls", "detection controls",
    "detection control", "recommended action", "responsibility", "action result",
    "classification", "special characteristic", "potential causes", "mechanism",
    "check item",
})

# Low-signal keywords that appear in data rows — penalise when dominant
_WEAK_KW: frozenset[str] = frozenset({
    "internal", "external", "customer", "confidential", "draft", "approved",
    "open", "closed", "tbd", "n/a", "yes", "no", "high", "medium", "low",
    "product", "charger", "scope", "boundary", "scope boundary",
    "confidentiality",
})


def build_strong_kw_set() -> frozenset[str]:
    """Return the shared strong-keyword set (exported for tests and extensions)."""
    return _STRONG_KW


# ---------------------------------------------------------------------------
# Header-row detection
# ---------------------------------------------------------------------------

def detect_template_columns(
    xlsx_path: Path,
    max_scan_rows: int = 60,
    min_columns: int = 2,
) -> list[str]:
    """
    Scan the active worksheet of an .xlsx / .xlsm file and return the column
    names from the row most likely to be the header row.

    The scoring heuristic:
      - Each cell value normalised to lowercase alphanumeric + spaces.
      - strong_hits  × 5  (strong engineering keywords)
      - + len(vals)       (more columns = more likely a header)
      - - weak_hits  × 3  (data-row indicators penalised)
      - A row must have at least 1 strong hit to qualify.

    Returns an empty list if no suitable header row is found or on any error.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        best_cols: list[str] = []
        best_score = -1

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 1, max_scan_rows),
            values_only=True,
        ):
            vals = [str(x).strip() for x in row if x not in (None, "")]
            if len(vals) < min_columns:
                continue

            norm = [
                re.sub(r"[^a-z0-9 /]+", "", v.lower()).strip()
                for v in vals
            ]
            strong_hits = sum(1 for n in norm if n in _STRONG_KW)
            weak_hits   = sum(1 for n in norm if n in _WEAK_KW)

            # Must have at least one strong keyword to be a real header
            if strong_hits == 0:
                continue

            score = strong_hits * 5 + len(vals) - weak_hits * 3
            if score > best_score:
                best_score = score
                best_cols = vals

        wb.close()
        return best_cols if len(best_cols) >= min_columns else []

    except Exception:
        return []


# ---------------------------------------------------------------------------
# File upload helper
# ---------------------------------------------------------------------------

async def save_upload(
    file: UploadFile,
    sid: str,
    uploads_dir: Path,
    max_upload_mb: int = 25,
    allowed_suffixes: frozenset[str] | None = None,
) -> tuple[Path, str]:
    """
    Read an uploaded file, enforce size limit, and write it to uploads_dir.

    Parameters
    ----------
    file             : The FastAPI UploadFile object.
    sid              : Session ID (used to namespace the saved filename).
    uploads_dir      : Directory to write the file into (created if needed).
    max_upload_mb    : Maximum allowed file size in megabytes.
    allowed_suffixes : Optional set of allowed file extensions, e.g.
                       frozenset({".pdf", ".xlsx"}).  Pass None to allow all.

    Returns
    -------
    (dest_path, original_filename) where dest_path is the Path of the saved
    file on disk.

    Raises
    ------
    HTTPException 400 if the file suffix is not allowed.
    HTTPException 413 if the file exceeds max_upload_mb.
    """
    name = Path(file.filename or "upload.bin").name
    suffix = Path(name).suffix.lower()

    if allowed_suffixes is not None and suffix not in allowed_suffixes:
        raise HTTPException(
            400,
            f"File type '{suffix}' is not accepted. "
            f"Allowed: {', '.join(sorted(allowed_suffixes))}",
        )

    data = await file.read()

    if len(data) > max_upload_mb * 1024 * 1024:
        raise HTTPException(413, f"File exceeds the {max_upload_mb} MB limit.")

    uploads_dir.mkdir(parents=True, exist_ok=True)
    dest = uploads_dir / f"{sid}_{uuid4().hex[:8]}_{name}"
    dest.write_bytes(data)

    return dest, name

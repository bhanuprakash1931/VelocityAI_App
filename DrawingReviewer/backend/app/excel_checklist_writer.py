"""Excel checklist writer for Drawing Reviewer — ported from VelocityAI_Platform."""
from __future__ import annotations

import html
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_DIR = Path("data/generated_checklists")

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("id", "sl", "s no", "s.no", "no", "item no", "finding id"),
    "check_item": ("check item", "check point", "checkpoint", "checklist", "requirement", "criteria", "item", "description", "review item", "question"),
    "answer": ("yes/no/na", "yes no na", "yes/no", "y/n", "answer", "response", "compliance", "compliant", "pass/fail", "pass fail", "result", "status", "outcome"),
    "comment": ("comment", "comments", "justification", "evidence", "remarks", "remark", "observation", "observations", "notes", "explanation"),
    "finding": ("finding", "findings", "issue", "issues", "non conformance", "gap"),
    "recommendation": ("recommendation", "recommendations", "corrective action", "action", "suggestion", "mitigation", "next step"),
    "severity": ("severity", "priority", "risk", "criticality"),
    "category": ("category", "area", "section", "discipline"),
    "checked": ("checked", "reviewed", "done", "completed"),
}

SEVERITY_RANK = {"critical": 4, "error": 3, "warning": 2, "info": 1, "": 0}


def _safe_stem(name: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in name).strip("_") or "drawing_review_checklist"


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9/+.-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _match_header(value: Any) -> str | None:
    header = _norm(value)
    if not header:
        return None
    # Exact matches first
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if header == _norm(alias):
                return field
    # Contains matches second (ignore short aliases)
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _norm(alias)
            if len(alias_norm) >= 3 and alias_norm in header:
                return field
    return None


def _find_header_row(ws) -> tuple[int | None, dict[str, int]]:
    best_row = None
    best_map: dict[str, int] = {}
    best_score = 0
    for row in range(1, min(ws.max_row, 50) + 1):
        col_map: dict[str, int] = {}
        for col in range(1, min(ws.max_column, 60) + 1):
            field = _match_header(ws.cell(row, col).value)
            if field and field not in col_map:
                col_map[field] = col
        score = len(col_map)
        if "check_item" in col_map:
            score += 3
        if "answer" in col_map:
            score += 2
        if "comment" in col_map:
            score += 1
        if score > best_score:
            best_score = score
            best_row = row
            best_map = col_map
    if best_score >= 3:
        return best_row, best_map
    return None, {}


def _choose_sheet(wb) -> tuple[Any, int | None, dict[str, int]]:
    best = None
    best_score = -1
    for ws in wb.worksheets:
        header_row, col_map = _find_header_row(ws)
        score = len(col_map)
        name = ws.title.lower()
        if any(key in name for key in ("check", "review", "drawing", "inspection")):
            score += 2
        if header_row and score > best_score:
            best = (ws, header_row, col_map)
            best_score = score
    if best:
        return best
    return wb.active, None, {}


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    if src_row < 1 or src_row == dst_row:
        return
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)
        dst.border = copy(src.border)


def _row_text(ws, row: int, max_col: int) -> str:
    return " | ".join(_clean(ws.cell(row, col).value) for col in range(1, max_col + 1) if _clean(ws.cell(row, col).value))


def _data_rows(ws, header_row: int, col_map: dict[str, int]) -> list[int]:
    key_cols = [col_map[key] for key in ("check_item", "category", "answer", "comment") if key in col_map]
    if not key_cols:
        key_cols = list(col_map.values())
    rows = []
    blank_streak = 0
    for row in range(header_row + 1, ws.max_row + 1):
        has_value = any(_clean(ws.cell(row, col).value) for col in key_cols)
        if has_value:
            rows.append(row)
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 8:
                break
    return rows


def _finding_summary(check_results: dict[str, Any], checklist: list[dict[str, Any]]) -> list[dict[str, Any]]:
    raw = checklist or (check_results or {}).get("findings") or []
    out = []
    for i, item in enumerate(raw, 1):
        if not isinstance(item, dict):
            continue
        text = _clean(item.get("text") or item.get("description"))
        if not text:
            continue
        out.append({
            "id": item.get("id") or f"F{i}",
            "category": item.get("category") or "General",
            "severity": str(item.get("severity") or "info").lower(),
            "text": text,
            "recommendation": item.get("recommendation") or "",
        })
    return out


def _keywords(text: str) -> set[str]:
    stop = {
        "the", "and", "for", "with", "from", "this", "that", "drawing", "review", "check",
        "item", "shall", "must", "should", "verify", "ensure", "confirm", "available", "provided",
    }
    words = set(re.findall(r"[a-zA-Z0-9]{3,}", text.lower()))
    return {word for word in words if word not in stop}


def _match_findings(row_text: str, findings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    row_words = _keywords(row_text)
    row_norm = _norm(row_text)
    matched: list[dict[str, Any]] = []
    for finding in findings:
        haystack = " ".join([
            str(finding.get("id") or ""),
            str(finding.get("category") or ""),
            str(finding.get("text") or ""),
            str(finding.get("recommendation") or ""),
        ])
        category_norm = _norm(finding.get("category") or "")
        hay_words = _keywords(haystack)
        if category_norm and category_norm in row_norm:
            matched.append(finding)
        elif row_words and len(row_words.intersection(hay_words)) >= 2:
            matched.append(finding)
    return matched


def _evaluate_row(row_number: int, row_text: str, findings: list[dict[str, Any]]) -> dict[str, Any]:
    matches = _match_findings(row_text, findings)
    if matches:
        severity = max((str(m.get("severity") or "info") for m in matches), key=lambda s: SEVERITY_RANK.get(s.lower(), 0))
        comment = " ".join(f"{m.get('id')}: {m.get('text')}" for m in matches[:3])
        recommendation = " ".join(str(m.get("recommendation") or "") for m in matches[:3]).strip()
        return {"row": row_number, "answer": "No", "comment": comment, "finding": comment, "recommendation": recommendation, "severity": severity}
    lower = row_text.lower()
    if any(term in lower for term in ("not applicable", "n/a", "if applicable", "where applicable")):
        return {"row": row_number, "answer": "NA", "comment": "No related AI finding was identified; item appears conditional or not applicable from the checklist wording. Human reviewer should confirm.", "finding": "", "recommendation": "", "severity": ""}
    return {"row": row_number, "answer": "Yes", "comment": "No AI finding was identified against this checklist item. Human reviewer should confirm before release.", "finding": "", "recommendation": "", "severity": ""}


def _write_cell(ws, row: int, col: int | None, value: Any, *, wrap: bool = True) -> None:
    if not col:
        return
    cell = ws.cell(row, col)
    cell.value = value
    if wrap:
        try:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or "top",
                text_rotation=cell.alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=cell.alignment.shrink_to_fit,
                indent=cell.alignment.indent,
            )
        except Exception:
            pass


def _unique_sheet_name(wb, base_name: str) -> str:
    if base_name not in wb.sheetnames:
        return base_name
    for idx in range(2, 100):
        candidate = f"{base_name} {idx}"
        if candidate not in wb.sheetnames:
            return candidate
    return f"{base_name} Copy"


def _disable_workbook_protection(wb) -> None:
    try:
        wb.security.lockStructure = False
        wb.security.lockWindows = False
        wb.security.workbookPassword = None
    except Exception:
        pass
    for ws in wb.worksheets:
        try:
            ws.protection.sheet = False
            ws.protection.enable = False
        except Exception:
            pass


def _append_finding_row(ws, *, row_number: int, col_map: dict[str, int], finding: dict[str, Any], report_docx_path: str | None = None) -> None:
    _write_cell(ws, row_number, col_map.get("id"), finding.get("id") or f"F{row_number}")
    _write_cell(ws, row_number, col_map.get("category"), finding.get("category") or "AI Finding")
    _write_cell(ws, row_number, col_map.get("check_item"), finding.get("category") or finding.get("text") or "AI Finding")
    _write_cell(ws, row_number, col_map.get("answer"), "No")
    _write_cell(ws, row_number, col_map.get("comment"), finding.get("text") or "")
    _write_cell(ws, row_number, col_map.get("finding"), finding.get("text") or "")
    _write_cell(ws, row_number, col_map.get("recommendation"), finding.get("recommendation") or "")
    _write_cell(ws, row_number, col_map.get("severity"), finding.get("severity") or "info")
    _write_cell(ws, row_number, col_map.get("checked"), True, wrap=False)
    try:
        severity_col = col_map.get("severity")
        if severity_col:
            severity = str(finding.get("severity") or "info").lower()
            fill = {"critical": "F4CCCC", "error": "FCE4D6", "warning": "FFF2CC", "info": "DDEBF7"}.get(severity, "DDEBF7")
            ws.cell(row_number, severity_col).fill = PatternFill("solid", fgColor=fill)
    except Exception:
        pass
    if report_docx_path:
        try:
            target_col = col_map.get("comment") or col_map.get("finding") or col_map.get("recommendation") or 1
            cell = ws.cell(row_number, target_col)
            cell.comment = Comment(f"Generated Word report: {report_docx_path}", "Velocity AI")
        except Exception:
            pass


def _add_ai_review_summary_sheet(wb, *, findings: list[dict[str, Any]], extracted_data: dict[str, Any] | None, report_docx_path: str | None) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, "AI Review Summary"))
    headers = ["Section", "Field", "Value"]
    ws.append(headers)
    for col in range(1, 4):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row = 2
    title_block = (extracted_data or {}).get("title_block") or {}
    if isinstance(title_block, dict):
        for key, value in title_block.items():
            ws.cell(row, 1).value = "Title Block"
            ws.cell(row, 2).value = str(key)
            ws.cell(row, 3).value = _clean(value)
            row += 1
    elif title_block:
        ws.cell(row, 1).value = "Title Block"
        ws.cell(row, 2).value = "Extracted"
        ws.cell(row, 3).value = _clean(title_block)
        row += 1
    if report_docx_path:
        ws.cell(row, 1).value = "Generated Artifact"
        ws.cell(row, 2).value = "Word Report"
        ws.cell(row, 3).value = report_docx_path
        try:
            ws.cell(row, 3).hyperlink = report_docx_path
            ws.cell(row, 3).style = "Hyperlink"
        except Exception:
            pass
        row += 1
    row += 1
    finding_header_row = row
    finding_headers = ["ID", "Severity", "Category", "Finding", "Recommendation", "Checked"]
    for idx, header in enumerate(finding_headers, 1):
        cell = ws.cell(finding_header_row, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
    if findings:
        for finding in findings:
            ws.cell(row, 1).value = finding.get("id")
            ws.cell(row, 2).value = finding.get("severity")
            ws.cell(row, 3).value = finding.get("category")
            ws.cell(row, 4).value = finding.get("text")
            ws.cell(row, 5).value = finding.get("recommendation")
            ws.cell(row, 6).value = bool(finding.get("checked", False))
            row += 1
    else:
        ws.cell(row, 1).value = "-"
        ws.cell(row, 2).value = "info"
        ws.cell(row, 3).value = "General"
        ws.cell(row, 4).value = "No AI findings were generated. Human reviewer should confirm."
        ws.cell(row, 6).value = False
    widths = [18, 18, 26, 80, 80, 14]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            try:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            except Exception:
                pass
    try:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass


def _create_fallback_sheet(wb, findings: list[dict[str, Any]], report_docx_path: str | None) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, "AI Filled Checklist"))
    ws.append(["Checklist Item", "Answer", "Comment", "Finding", "Recommendation", "Severity", "Report Document"])
    if findings:
        for finding in findings:
            ws.append([finding.get("category") or "AI Finding", "No", finding.get("text"), finding.get("text"), finding.get("recommendation"), finding.get("severity"), report_docx_path or ""])
    else:
        ws.append(["Drawing review", "Yes", "No AI findings were generated. Human reviewer should confirm.", "", "", "", report_docx_path or ""])
    for idx, width in enumerate([28, 12, 80, 80, 80, 14, 80], 1):
        ws.column_dimensions[chr(64 + idx)].width = width


def fill_uploaded_checklist_copy_fast(
    *,
    template_path: str,
    checklist: list[dict[str, Any]],
    extracted_data: dict[str, Any] | None = None,
    check_results: dict[str, Any] | None = None,
    report_docx_path: str | None = None,
) -> str:
    source = Path(template_path)
    if not source.exists():
        raise FileNotFoundError(f"Checklist template not found: {template_path}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Checklist update supports .xlsx and .xlsm files.")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"{_safe_stem(source.stem)}_filled_checklist_EDITABLE_{timestamp}{source.suffix.lower()}"
    shutil.copy2(source, output_path)

    wb = load_workbook(output_path, keep_vba=source.suffix.lower() == ".xlsm")
    _disable_workbook_protection(wb)

    ws, header_row, col_map = _choose_sheet(wb)
    findings = _finding_summary(check_results or {}, checklist or [])
    consumed_finding_ids: set[str] = set()

    if header_row and col_map:
        data_rows = _data_rows(ws, header_row, col_map)
        for row_number in data_rows:
            row_text = _row_text(ws, row_number, min(ws.max_column, 60))
            if not row_text:
                continue
            matched = _match_findings(row_text, findings)
            for finding in matched:
                if finding.get("id"):
                    consumed_finding_ids.add(str(finding.get("id")))
            ev = _evaluate_row(row_number, row_text, findings)
            _write_cell(ws, row_number, col_map.get("answer"), ev["answer"])
            _write_cell(ws, row_number, col_map.get("comment"), ev["comment"])
            _write_cell(ws, row_number, col_map.get("finding"), ev["finding"])
            _write_cell(ws, row_number, col_map.get("recommendation"), ev["recommendation"])
            _write_cell(ws, row_number, col_map.get("severity"), ev["severity"])
            _write_cell(ws, row_number, col_map.get("checked"), True, wrap=False)

        unmatched = [f for f in findings if str(f.get("id") or "") not in consumed_finding_ids]
        append_start = (data_rows[-1] + 1) if data_rows else (header_row + 1)
        max_col = max(ws.max_column, max(col_map.values()) if col_map else 1)
        style_source_row = data_rows[-1] if data_rows else header_row
        for offset, finding in enumerate(unmatched):
            row_number = append_start + offset
            _copy_row_style(ws, style_source_row, row_number, max_col)
            _append_finding_row(ws, row_number=row_number, col_map=col_map, finding=finding, report_docx_path=report_docx_path)

        if report_docx_path:
            note_row = max(ws.max_row + 2, (header_row or 1) + 2)
            ws.cell(note_row, 1).value = "Generated Word Report"
            ws.cell(note_row, 2).value = report_docx_path
            try:
                ws.cell(note_row, 2).hyperlink = report_docx_path
                ws.cell(note_row, 2).style = "Hyperlink"
            except Exception:
                pass
    else:
        _create_fallback_sheet(wb, findings, report_docx_path)

    _add_ai_review_summary_sheet(wb, findings=findings, extracted_data=extracted_data or {}, report_docx_path=report_docx_path)
    _disable_workbook_protection(wb)
    wb.save(output_path)
    return str(output_path)


def build_filled_checklist_preview_html(file_path: str, max_rows: int = 30) -> str:
    path = Path(file_path)
    if not path.exists():
        return "<h1>Filled Checklist</h1><p>No filled checklist was generated.</p>"
    wb = load_workbook(path, data_only=True)
    ws, header_row, col_map = _choose_sheet(wb)
    rows_html = ""
    if header_row and col_map:
        display_cols = []
        for key in ("id", "check_item", "answer", "comment", "finding", "recommendation", "severity"):
            if key in col_map:
                display_cols.append((key, col_map[key]))
        headers = "".join(f"<th>{html.escape(key.replace('_', ' ').title())}</th>" for key, _ in display_cols)
        for row_number in _data_rows(ws, header_row, col_map)[:max_rows]:
            cells = "".join(f"<td>{html.escape(_clean(ws.cell(row_number, col).value))}</td>" for _, col in display_cols)
            rows_html += f"<tr>{cells}</tr>"
        table = f"<table border='1' cellspacing='0' cellpadding='6' style='border-collapse:collapse;width:100%;'><thead><tr style='background:#1F4E78;color:white;'>{headers}</tr></thead><tbody>{rows_html}</tbody></table>"
    else:
        ws = wb.active
        max_col = min(ws.max_column, 8)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), max_col=max_col, values_only=True):
            rows_html += "<tr>" + "".join(f"<td>{html.escape(_clean(value))}</td>" for value in row) + "</tr>"
        table = f"<table border='1' cellspacing='0' cellpadding='6' style='border-collapse:collapse;width:100%;'><tbody>{rows_html}</tbody></table>"
    return f"""
    <div style="font-family: Aptos, Arial, sans-serif; font-size: 13px; line-height: 1.45; padding: 14px;">
      <h1>Filled Drawing Review Checklist</h1>
      <p><b>Generated file:</b> {html.escape(path.name)}</p>
      <p>The uploaded checklist template was filled as a copied workbook. The original uploaded file was not modified.</p>
      {table}
    </div>
    """


# Backward-compatible aliases
def fill_uploaded_checklist_copy(*, template_path: str, checklist: list[dict[str, Any]], report_docx_path: str | None = None) -> str:
    return fill_uploaded_checklist_copy_fast(template_path=template_path, checklist=checklist, extracted_data={}, check_results={}, report_docx_path=report_docx_path)


def update_uploaded_checklist_workbook(*, template_path: str, checklist: list[dict[str, Any]], report_docx_path: str | None = None) -> str:
    return fill_uploaded_checklist_copy(template_path=template_path, checklist=checklist, report_docx_path=report_docx_path)